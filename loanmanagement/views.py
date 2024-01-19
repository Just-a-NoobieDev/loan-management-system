import datetime
import json
import mimetypes
import os
from decimal import Decimal
from wsgiref.util import FileWrapper

from django.conf import settings
from django.contrib import auth, messages
from django.contrib.auth import logout, authenticate, login
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.shortcuts import render, HttpResponse, redirect, get_object_or_404
from openpyxl.workbook import Workbook
from qrcode import *
from django.db.models import Count, Sum
from .forms import LoanForm, PaymentForm
from .forms import PersonForm, UserRegistration, CollectorForm
from .models import Loan, Payment, Reports
from .models import Person, Collector
from django.http import HttpResponse
from io import BytesIO
import base64
from django.db.models.functions import TruncMonth
import matplotlib.pyplot as plt
from django.db import transaction

# Create your views here.


def clients_with_active_loans_count():
    count = (
        Loan.objects.filter(has_active_loan=True)
        .values("client_id")
        .annotate(client_count=Count("client_id", distinct=True))
        .count()
    )

    return count


def sum_of_loan_amounts():
    sum_amount = Loan.objects.aggregate(Sum("loan_amount"))["loan_amount__sum"]

    return sum_amount or 0


def sum_of_payments():
    sum_payment = Payment.objects.aggregate(Sum("amount"))["amount__sum"]
    return sum_payment or 0


def home(request):
    user = request.user
    if not user.is_authenticated:
        return redirect("adminLogin")
    active_clients_count = clients_with_active_loans_count()
    total_loan_amount = sum_of_loan_amounts()
    payment_amount = sum_of_payments()
    formatted_total_loan_amount = f"₱ {total_loan_amount:,.2f}"
    formatted_payment_amount = f"₱ {payment_amount:,.2f}"
    p = Person.objects.all().count()
    latest_persons = Person.objects.order_by("-created_at")[:5]
    recent_payments = Payment.objects.select_related("loan_id").order_by(
        "-payment_date"
    )[:5]
    monthly_loan_amounts = (
        Loan.objects.annotate(month=TruncMonth("loan_date"))
        .values("month")
        .annotate(total_amount=Sum("loan_amount"))
    )

    # Extract data for Matplotlib chart
    months = [entry["month"].strftime("%b %Y") for entry in monthly_loan_amounts]
    total_amounts = [entry["total_amount"] for entry in monthly_loan_amounts]

    plt.figure(figsize=(8, 5))
    plt.bar(months, total_amounts, color="#09090b", edgecolor="black", linewidth=0.8)
    # Inside the loop for adding annotations

    for i, count in enumerate(total_amounts):
        formatted_count = f"₱ {count:,.2f}"  # Format count as currency
        plt.text(
            i,
            count / 2,
            formatted_count,
            ha="center",
            va="center",
            fontsize=10,
            color="white",
        )

    plt.xlabel("Month", fontsize=12, color="white")
    plt.xticks(rotation=45, ha="right", fontsize=12, color="black")
    plt.yticks(fontsize=12, color="black")
    plt.grid(axis="y", linestyle="-", alpha=0)
    plt.gca().set_facecolor("#f8f8fd")

    image_stream = BytesIO()
    plt.savefig(image_stream, format="png", bbox_inches="tight", facecolor="#f8f8fd")
    plt.close()

    image_base64 = base64.b64encode(image_stream.getvalue()).decode("utf-8")
    return render(
        request,
        "dashboard.html",
        {
            "p": p,
            "active_clients_count": active_clients_count,
            "total_loan_amount": formatted_total_loan_amount,
            "payment_amount": formatted_payment_amount,
            "image_base64": image_base64,
            "latest_persons": latest_persons,
            "recent_payments": recent_payments,
        },
    )


def addClient(request):
    loans = Person.objects.all()
    paginated = Paginator(loans, 10)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)
    user = request.user
    if not user.is_authenticated:
        return redirect("adminLogin")
    if request.method == "POST":
        form = PersonForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        else:
            context = {"form": form}
            return render(request, "addClient.html", context)

    persons = Person.objects.all()
    context = {"form": PersonForm(), "persons": page, "paginator": paginated}
    return render(request, "addClient.html", context)


def deleteClient(request, id):
    item = Person.objects.get(id=id)
    item.delete()
    if item.picture.name and item.document.name:
        os.remove(os.path.join(settings.MEDIA_ROOT, item.picture.name))
        os.remove(os.path.join(settings.MEDIA_ROOT, item.document.name))
    return redirect("/adminUser/addClient")


def qrFunc(val, name):
    img = make(val)
    img_url = "qr-" + str(name) + ".png"
    # buffer = io.BytesIO()
    img.save(settings.MEDIA_ROOT + "/qr/" + img_url)
    return img_url


def generateQR(request, id):
    person = Person.objects.get(id=id)
    img = qrFunc(person.id, person.name)
    # # qrcode = base64.b64encode(buffer.getvalue()).decode("utf-8")
    if img:
        wrapper = FileWrapper(open(settings.MEDIA_ROOT + "/qr/" + img, "rb"))
        content_type = mimetypes.guess_type(img)[0]
        response = HttpResponse(wrapper, content_type=content_type)
        response["Content-Disposition"] = "attachment; filename=%s" % img
        return response


def loan_list(request):
    user = request.user
    if not user.is_authenticated:
        return redirect("adminLogin")
    loans = Loan.objects.all()
    paginated = Paginator(loans, 10)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)

    if request.method == "POST":
        form = LoanForm(request.POST)
        if form.is_valid():
            loan = form.save(commit=False)

            if loan.loan_balance == 0:
                loan.has_active_loan = False
            loan.save()
    else:
        form = LoanForm()
    return render(
        request,
        "loanPortal.html",
        {"loans": page, "form": form, "paginator": paginated},
    )


def add_loan(request):
    if request.method == "POST":
        form = LoanForm(request.POST)
        if form.is_valid():
            loan = form.save(commit=False)

            existing_active_loan = Loan.objects.filter(
                client_id=loan.client_id, has_active_loan=True
            ).exists()

            if existing_active_loan:
                messages.error(request, "Client already has an active loan.")
                return JsonResponse({"status": "error", "redirect": "loanList/"})
            else:
                # Calculate and update net, loan_balance
                net_percentage = (
                    (loan.interest_rate / 100) * loan.loan_amount
                ) + loan.processing_fee
                loan.loan_balance += net_percentage
                loan.save()

                messages.success(request, "Loan added successfully.")
                return JsonResponse({"status": "success", "redirect": "loanList/"})
        else:
            messages.error(
                request, "Error in the form submission. Please check the form."
            )
            return JsonResponse({"status": "error", "redirect": "loanList/"})
    else:
        form = LoanForm()

    return render(request, "loanPortal.html", {"form": form})


# def delete_loan(request, id):
#     itemCol = Loan.objects.get(id=id)
#     itemCol.delete()
#     messages.success(request, 'Loan deleted successfully.')
#     return redirect('/adminUser/loanList/')


def collectorLogin(request):
    return render(request, "collectorLogin.html")


# def loanPortal(request):
#     return render(request, "loanPortal.html")


def paymentList(request):
    user = request.user
    if not user.is_authenticated:
        return redirect("adminLogin")
    payments = Payment.objects.all()
    paginated = Paginator(payments, 15)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)
    form = PaymentForm(request.POST)

    return render(
        request,
        "payments.html",
        {"payments": page, "form": form, "paginator": paginated},
    )


def addPayment(request):
    if request.method == "POST":
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            loan = payment.loan_id
            loan.loan_balance -= payment.amount
            loan.save()
            payment.save()
            messages.success(request, "Payment added successfully.")
            return JsonResponse({"status": "success", "redirect": "paymentList/"})
        else:
            messages.error(
                request, "Error in the form submission. Please check the form."
            )
            return JsonResponse({"status": "error", "redirect": "paymentList/"})
    else:
        form = PaymentForm()

    return render(request, "payments.html", {"form": form})


def edit_payment(request, id):
    # Assuming Payment is your model
    payment_instance = get_object_or_404(Payment, id=id)

    # Set the amount to 200
    payment_instance.amount = 200

    # Save the changes
    payment_instance.save()

    messages.success(request, "Payment edited successfully.")
    return redirect("/adminUser/paymentList/")


def delete_payment(request, id):
    itemCol = Payment.objects.get(id=id)
    amount = itemCol.amount
    ll = Loan.objects.get(id=itemCol.loan_id.id)
    ll.loan_balance = ll.loan_balance + amount
    ll.save()
    itemCol.delete()
    messages.success(request, "Payment deleted successfully.")
    return redirect("/adminUser/paymentList/")


def editPayment(request):
    loan_id = request.POST.get("loan_id")
    ciId = int(request.POST.get("id"))
    camount = Decimal(request.POST.get("camount"))
    namount = Decimal(request.POST.get("namount"))

    c = Payment.objects.get(id=ciId)
    c.amount = namount
    c.save()

    l = Loan.objects.get(id=loan_id)
    l.loan_balance = l.loan_balance + camount

    l.loan_balance = l.loan_balance - namount
    if l.loan_balance > 0:
        l.has_active_loan = True

    l.save()

    return redirect("/adminUser/paymentList/")


def payment(request):
    id = request.POST.get("id")
    payment_data = Payment.objects.get(id=id)
    payment_json = {
        "amount": payment_data.amount,
    }
    return HttpResponse(json.dumps(payment_json), content_type="application/json")


# def payments(request):
#     posts = Payment.objects.all()
#     paginated = Paginator(posts, 3)
#     page_number = request.GET.get('page') #Get the requested page number from the URL

#     page = paginated.get_page(page_number)
#     return render(request, 'adminUser/paymentList', {'page':page})


def reports(request):
    reports = Reports.objects.all()
    paginated = Paginator(reports, 10)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)
    user = request.user
    if not user.is_authenticated:
        return redirect("adminLogin")
    report = Reports.objects.all()
    return render(request, "reports.html", {"reports": page, "paginator": paginated})


def register(request):
    user = request.user
    if user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        form = UserRegistration(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
        # else:
        #     return render(request, 'register.html', {'form': form})
        #     raise Exception(form.errors.get_text())
    else:
        form = UserRegistration()
    return render(request, "register.html", {"form": form})


def adminLogin(request):
    user = request.user
    if user.is_authenticated:
        return redirect("dashboard")
    return render(request, "adminLogin.html")


def loginController(request):
    logout(request)
    resp = {"status": "failed", "msg": ""}
    username = ""
    password = ""
    if request.POST:
        username = request.POST["user"]
        password = request.POST["password"]

        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                resp["status"] = "success"
            else:
                resp["msg"] = "Incorrect username or password"
        else:
            resp["msg"] = "Incorrect username or password"
    return HttpResponse(json.dumps(resp), content_type="application/json")


def logout_page(request):
    auth.logout(request)
    return redirect("/adminUser/adminLogin/")


def client(request):
    id = request.POST.get("id")
    client_data = Person.objects.get(id=id)
    client_json = {
        "name": client_data.name,
        "address": client_data.address,
        "email": client_data.email,
        "password": client_data.password,
        "image": "/media/" + client_data.picture.name,
        "document": "/media/" + client_data.document.name,
    }

    print(client_json)
    return HttpResponse(json.dumps(client_json), content_type="application/json")


def addCollector(request):
    persons = Collector.objects.all()
    paginated = Paginator(persons, 10)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)
    user = request.user
    if not user.is_authenticated:
        return redirect("adminLogin")
    if request.method == "POST":
        colForm = CollectorForm(request.POST)
        if colForm.is_valid():
            colForm.save()
        else:
            context = {"form": colForm}
            return render(request, "addClient.html", context)

    collector = Collector.objects.all()
    colForm = CollectorForm(request.POST)
    return render(
        request,
        "addCollector.html",
        {"persons": page, "form": colForm, "paginator": paginated},
    )


def generateCollector(request, id):
    col = Collector.objects.get(id=id)
    img = qrFunc(col.id, col.name)
    # # qrcode = base64.b64encode(buffer.getvalue()).decode("utf-8")
    if img:
        wrapper = FileWrapper(open(settings.MEDIA_ROOT + "/qr/" + img, "rb"))
        content_type = mimetypes.guess_type(img)[0]
        response = HttpResponse(wrapper, content_type=content_type)
        response["Content-Disposition"] = "attachment; filename=%s" % img
        return response


def deleteCollector(request, id):
    itemCol = Collector.objects.get(id=id)
    itemCol.delete()
    return redirect("/adminUser/addCollector")


def singleCollector(request):
    collectorId = request.POST.get("id")
    collector_data = Collector.objects.get(id=collectorId)
    collector_json = {
        "name": collector_data.name,
        "email": collector_data.email,
        "pass": collector_data.password,
    }
    return HttpResponse(json.dumps(collector_json), content_type="application/json")


def editClient(request):
    cId = request.POST.get("id")
    name = request.POST.get("name")
    address = request.POST.get("address")
    email = request.POST.get("email")
    password = request.POST.get("password")

    c = Person.objects.get(id=cId)
    c.name = name
    c.address = address
    c.email = email
    c.password = password
    c.save()
    return redirect("/adminUser/addClient")


def editCollector(request):
    colId = request.POST.get("id")
    name = request.POST.get("name")
    email = request.POST.get("email")
    password = request.POST.get("password")

    col = Collector.objects.get(id=colId)
    col.name = name
    col.email = email
    col.password = password
    col.save()
    return redirect("/adminUser/addCollector")


def exportTodayPayments(request):
    ptfilename = f"Payments Today - {datetime.date.today()}.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{ptfilename}"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Today Payments"

    # Add headers
    headers = ["ID", "Borrower Name", "Amount", "Payment Date", "OR Number"]
    ws.append(headers)

    # Add data from the model
    payments = Payment.objects.filter(payment_date__exact=datetime.date.today())
    for payment in payments:
        ws.append(
            [
                payment.id,
                payment.loan_id.client_id.name,
                payment.amount,
                payment.payment_date.strftime("%m/%d/%Y"),
                payment.or_number,
            ]
        )

    # Save the workbook to the HttpResponse
    path = settings.MEDIA_ROOT + "/reports/payments/" + ptfilename
    wb.save(response)
    wb.save(path)

    todayReport = Reports(type="Payments", filename=ptfilename, filepath=path)
    todayReport.save()

    return response


def exportTodayLoans(request):
    ltfilename = f"Loans Today - {datetime.date.today()}.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{ltfilename}"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Today Loans"

    # Add headers
    headers = [
        "ID",
        "Borrower Name",
        "Loan Date",
        "Duration Period",
        "Interest Rate",
        "Loan Amount",
        "Loan Maturity",
        "Guarantor",
        "Processing Fee",
        "Net",
        "Checking Number",
    ]
    ws.append(headers)

    # Add data from the model
    loans = Loan.objects.filter(loan_date__exact=datetime.date.today())
    for loan in loans:
        ws.append(
            [
                loan.id,
                loan.client_id.name,
                loan.loan_date.strftime("%m/%d/%Y"),
                loan.duration_period,
                loan.interest_rate,
                loan.loan_amount,
                loan.loan_maturity,
                loan.guarantor,
                loan.processing_fee,
                loan.net,
                loan.checking_no,
            ]
        )

    # Save the workbook to the HttpResponse
    path = settings.MEDIA_ROOT + "/reports/loans/" + ltfilename
    wb.save(response)
    wb.save(path)

    todayLReport = Reports(type="Loans", filename=ltfilename, filepath=path)
    todayLReport.save()

    return response


def exportLoans(request, dateFrom, dateTo):
    lfilename = f"Loans - {dateFrom} - {dateTo}.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{lfilename}"'

    wb = Workbook()
    ws = wb.active
    ws.title = f"Loans {dateFrom} - {dateTo}"

    # Add headers
    headers = [
        "ID",
        "Borrower Name",
        "Loan Date",
        "Duration Period",
        "Interest Rate",
        "Loan Amount",
        "Loan Maturity",
        "Guarantor",
        "Processing Fee",
        "Net",
        "Checking Number",
    ]
    ws.append(headers)

    # Add data from the model
    loans = Loan.objects.filter(loan_date__range=[dateFrom, dateTo])
    for loan in loans:
        ws.append(
            [
                loan.id,
                loan.client_id.name,
                loan.loan_date.strftime("%m/%d/%Y"),
                loan.duration_period,
                loan.interest_rate,
                loan.loan_amount,
                loan.loan_maturity,
                loan.guarantor,
                loan.processing_fee,
                loan.net,
                loan.checking_no,
            ]
        )

    # Save the workbook to the HttpResponse
    path = settings.MEDIA_ROOT + "/reports/loans/" + lfilename
    wb.save(response)
    wb.save(path)

    todayLReport = Reports(type="Loans", filename=lfilename, filepath=path)
    todayLReport.save()

    return response


def exportPayment(request, dateFrom, dateTo):
    pfilename = f"Payments - {dateFrom} - {dateTo}.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{pfilename}"'

    wb = Workbook()
    ws = wb.active
    ws.title = f"Payments {dateFrom} - {dateTo}"

    # Add headers
    headers = ["ID", "Borrower Name", "Amount", "Payment Date", "OR Number"]
    ws.append(headers)

    # Add data from the model
    payments = Payment.objects.filter(payment_date__range=[dateFrom, dateTo])
    for payment in payments:
        ws.append(
            [
                payment.id,
                payment.loan_id.client_id.name,
                payment.amount,
                payment.payment_date.strftime("%m/%d/%Y"),
                payment.or_number,
            ]
        )

    # Save the workbook to the HttpResponse
    path = settings.MEDIA_ROOT + "/reports/payments/" + pfilename
    wb.save(response)
    wb.save(path)

    todayReport = Reports(type="Payments", filename=pfilename, filepath=path)
    todayReport.save()

    return response


def logoutNow(request):
    logout(request)
    return redirect("adminLogin")
